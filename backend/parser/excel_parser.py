"""Excel (.xlsx / .xls) parsing utilities for the ingestion pipeline."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from ingestion.models import DocumentContent


def _row_to_human_readable(headers: list[str], row: list[Any]) -> str:
    """Convert a spreadsheet row into human-readable key-value format."""
    pairs: list[str] = []
    for header, value in zip(headers, row):
        # Normalise None and numeric types to strings
        str_value = "" if value is None else str(value)
        pairs.append(f"{header}: {str_value}")
    return "\n".join(pairs)


def parse_excel(*, excel_bytes: bytes, filename: str) -> DocumentContent:
    """Parse Excel bytes into a structured DocumentContent object.

    Each sheet is treated as a logical section. Within each sheet every data
    row is converted into a human-readable key-value block (same strategy as
    the CSV parser) so that downstream chunking, entity extraction, and
    embedding work correctly on tabular industrial data.
    """
    if not excel_bytes:
        raise ValueError("excel_bytes must not be empty")
    if not filename:
        raise ValueError("filename must not be empty")

    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to parse Excel files. "
            "Install it with: pip install openpyxl"
        ) from exc

    try:
        workbook = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"Unable to open Excel file: {exc}") from exc

    all_pages: list[str] = []
    total_rows = 0
    total_columns = 0
    sheet_names: list[str] = []

    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_names.append(sheet_name)

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            # First row is treated as the header
            raw_headers = rows[0]
            headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(raw_headers)]
            total_columns = max(total_columns, len(headers))

            data_rows = rows[1:]
            total_rows += len(data_rows)

            # Build a section header so chunking preserves sheet context
            section_header = f"=== Sheet: {sheet_name} ===\n"

            for row in data_rows:
                padded_row = list(row) + [None] * max(0, len(headers) - len(row))
                readable_text = _row_to_human_readable(headers, padded_row)
                all_pages.append(section_header + readable_text)
    finally:
        workbook.close()

    if not all_pages:
        raise ValueError("Excel file contains no data rows")

    text = "\n\n".join(all_pages)
    metadata: dict[str, str | int | bool | float] = {
        "sheet_names": ", ".join(sheet_names),
        "sheet_count": len(sheet_names),
        "row_count": total_rows,
        "column_count": total_columns,
        "source_format": "xlsx",
    }

    return DocumentContent(
        filename=filename,
        text=text,
        pages=tuple(all_pages),
        page_count=len(all_pages),
        metadata=metadata,
    )
